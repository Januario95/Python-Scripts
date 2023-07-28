# pip install aspose-cells

# from asposecells.api import Workbook

# Password protect Excel file
# wb.getSettings().setPassword("1234")

# # Encrypt by specifying the encryption type
# wb.setEncryptionOptions(EncryptionType.XOR, 40)

# # Specify Strong Encryption type (RC4,Microsoft Strong Cryptographic Provider)
# wb.setEncryptionOptions(EncryptionType.STRONG_CRYPTOGRAPHIC_PROVIDER, 128)

# # Save Excel file
# wb.save("workbook-encrypted.xlsx")










from openpyxl import Workbook


# wb = Workbook('janu-2.xlsx')
# wb.security.workbookPassword = "Jaci1995"
# wb.security.lockStructure = True
# # wb.save("janu-2.xlsx")



from openpyxl import load_workbook
from openpyxl.workbook.protection import WorkbookProtection

# wb = load_workbook('janu-2.xlsx')
# wb.security = WorkbookProtection(
#     workbookPassword='Jaci1995',
#     revisionsPassword='Jaci1995',
#     lockWindows=True,
#     lockStructure=True,
#     lockRevision=True)

# wb.save('janu-2.xlsx')
# wb.close()


# wb = Workbook()
# sheet1 = wb.active
# sheet1.cell(3, 3).value = "Januario"
# sheet1.cell(3, 4).value = "Cipriano"
# sheet1.protection.password = "Jaci1995"
# wb.save('Book1.xlsx')


# wb = load_workbook('janu-2.xlsx')
# sheet1 = wb.active
# sheet1.cell(3, 3).value = "Januario"
# sheet1.cell(3, 4).value = "Cipriano"
# sheet1.protection.password = "Jaci1995"
# wb.security = WorkbookProtection(
#     workbookPassword='Jaci1995',
#     revisionsPassword='Jaci1995',
#     lockWindows=True,
#     lockStructure=True,
#     lockRevision=True
# )
# wb.security.lock_windows = True
# wb.save('janu-3.xlsx')


import win32com.client as win32

# excel = win32.gencache.EnsureDispatch('Excel.Application')
# excel.DisplayAlerts = False
# wb = excel.Workbooks.Open('janu-3.xlsx')
# print(dir(wb))
# wb.SaveAs('janu-3.xlsx', 52, 'Jaci1995')
# wb.Visible = False
# wb.Close()
# excel.Application.Quit()




from win32com.client.gencache import EnsureDispatch
    
xlApp = EnsureDispatch("Excel.Application")
xlwb = xlApp.Workbooks.Open('janu-3.xlsx')
xlApp.DisplayAlerts = False
xlwb.Visible = False
xlwb.SaveAs('janu-3.xlsx', Password = 'Jaci1995')
xlwb.Close()
xlApp.Quit()
    
    












